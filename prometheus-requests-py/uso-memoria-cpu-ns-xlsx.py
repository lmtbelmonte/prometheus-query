# autor : Luis Merino Troncoso
# Fecha : 30/06/2023
#
# Descripcion : Extraccion de metricas(mem, cpu, requests, limits de Openshift a traves del REST Api de prometheus utilizando
#               utilizando libreria requests y openpyxl para generar libo excel          
# Ejecucion   : Hay que crear las variables de entorno despues de hacer el oc login al cluster
#               TOKEN: `oc whoami -t`
#               PROMETHEUS_URL: <route a prometheus-k8s>/api/v1/query/ 

import requests
import os
from openpyxl import Workbook

# cargamos la url de la ruta prometheus de la variable de entorno 
prometheus_url = os.environ['PROMETHEUS_URL']

# cargamos la url de la ruta prometheus de la variable de entorno 
token = os.environ['TOKEN']

# Preparamos la query de % maximo uso en bytes en un periodo de las requests / limits 
memory_query = 'sum(container_memory_usage_bytes{{namespace!="",container!="POD", container!=""}}) by (namespace,pod, container)'

# Preparamos la query de consumo CPU requests / limits 
cpu_query = 'sum(rate(container_cpu_usage_seconds_total{{namespace!="", container!="POD", container!=""}}[1h])) by (namespace,pod, container)'

# Recogemos parametros para llamada al api
params = {
    "query": ''
}

headers = {
    "Authorization": "Bearer "+token
}    

# Hacemos primero con la query de memoria la  request a Prometheus con la cabecera+token
params['query'] = memory_query
response = requests.get(prometheus_url, params=params, headers=headers, verify=False)
memory_data = response.json()

# Hacemos despues la query de CPU la  request a Prometheus con la cabecera+token
params['query'] = cpu_query
response = requests.get(prometheus_url, params=params, headers=headers, verify=False)
cpu_data = response.json()

# Extraemos del API response la informacion necesaria
memory_results = memory_data['data']['result']
cpu_results = cpu_data['data']['result']

# Creamos el workbook excel inicial 
workbook = Workbook()

# Creamos la worksheet para consumo memoria
memory_sheet = workbook.create_sheet('Consumo Memoria')
memory_sheet.append(['Namespace', 'Pod', 'Container','Uso Memoria'])

# Cargamos los datos consumo de memoria en la worksheet
for result in memory_results:
    namespace = result['metric']['namespace']
    pod = result['metric']['pod']
    container = result['metric']['container']
    memory_usage = result['value'][1]
    memory_sheet.append([namespace, pod, container, memory_usage])

# Creamos la worksheet de consumo  CPU 
cpu_sheet = workbook.create_sheet('Consumo CPU')
cpu_sheet.append(['Namespace', 'Pod', 'Container','Uso CPU'])

# Cargamos los datos de consumo CPU en la worksheet
for result in cpu_results:
    namespace = result['metric']['namespace']
    pod = result['metric']['pod']
    container = result['metric']['container']
    cpu_usage = result['value'][1]
    cpu_sheet.append([namespace, pod, container, cpu_usage])

# Grabamos la hoja excel 
workbook.save('/tmp/openshift_metrics.xlsx')
